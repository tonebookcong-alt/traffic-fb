#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cào dữ liệu Facebook Page (text + ảnh) bằng thư viện facebook-scraper.

Cách dùng (CLI):
    python scraper.py --page ten-trang --pages 5
    python scraper.py --page trangA trangB trangC --per-page 50
    python scraper.py --page ten-trang --cookies cookies.txt

Hoặc qua giao diện web:
    python webui.py   (mở http://127.0.0.1:5000)

Kết quả:
    - <output>.xlsx              : 1 file Excel chứa tất cả (sheet Tổng quan + 1 sheet/trang)
    - <output>_<trang>.json      : dữ liệu thô từng trang
    - du_lieu_images/<giờ-cào>_images/ : TẤT CẢ ảnh của lần cào (1 thư mục duy nhất
                                         theo giờ cào, nằm trong du_lieu_images/)
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime

import requests
try:
    from facebook_scraper import get_posts, set_cookies
except ImportError:
    get_posts = None
    set_cookies = None

import cao_fb  # engine cào mới: mở trình duyệt thật (Playwright + Edge/Chrome)

DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
)


def in_tu_page(page_arg: str) -> str:
    """Trích tên/ID trang từ URL hoặc tên trực tiếp.

    Xử lý cả link dạng profile.php?id=... (không bị cắt mất phần ?id=)."""
    m = re.search(r"facebook\.com/(profile\.php\?id=\d+)", page_arg)
    if m:
        return m.group(1)
    m = re.search(r"facebook\.com/([^/?]+)", page_arg)
    return m.group(1) if m else page_arg.strip("/").split("/")[-1]


def nghi_stop(giay: float, stop_flag=None):
    """Ngủ theo từng phần nhỏ (0.2s) để bấm Dừng là dừng được ngay,
    không phải đợi hết cả khoảng nghỉ."""
    if giay <= 0:
        return
    da_ngu = 0.0
    while da_ngu < giay:
        if stop_flag and stop_flag.is_set():
            return
        time.sleep(min(0.2, giay - da_ngu))
        da_ngu += 0.2


def load_cookies(path: str):
    """Đọc cookies.txt — hỗ trợ 3 định dạng:
    1. JSON (Cookie-Editor / Firefox / 'Get cookies.txt LOCALLY' dạng mới):
       [{"name": "...", "value": "...", "domain": "...", ...}]
    2. Netscape (extension cũ): tab-separated, 7 cột
    3. document.cookie từ DevTools: cac cap ten=giatri phan cach bang ';'
    """
    if not path or not os.path.isfile(path):
        return
    with open(path, "r", encoding="utf-8") as f:
        noi_dung = f.read().strip()
    cookies = {}

    # --- 1. Định dạng JSON ---
    if noi_dung.startswith("["):
        try:
            danh_sach = json.loads(noi_dung)
            if isinstance(danh_sach, list):
                for c in danh_sach:
                    if isinstance(c, dict) and c.get("name") and c.get("value"):
                        cookies[c["name"]] = c["value"]
        except json.JSONDecodeError:
            cookies = {}  # không phải JSON hợp lệ -> thử định dạng khác

    # --- 2. Netscape hoặc document.cookie ---
    if not cookies:
        for line in noi_dung.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) >= 7:  # dinh dang Netscape
                cookies[parts[5]] = parts[6]
            elif "=" in line:    # dinh dang document.cookie
                for cap in line.split(";"):
                    cap = cap.strip()
                    if "=" in cap:
                        k, v = cap.split("=", 1)
                        cookies[k.strip()] = v.strip()

    if cookies:
        if set_cookies:
            set_cookies(cookies)
        print(f"  [i] Đã nạp {len(cookies)} cookie từ {path}")
        # kiểm tra các cookie quan trọng để cào được
        thieu = [c for c in ("c_user", "xs", "datr") if c not in cookies]
        if thieu:
            print(f"  [!] Cảnh báo: thiếu cookie quan trọng: {', '.join(thieu)} "
                  f"— có thể vẫn không cào được")
        return cookies
    else:
        print(f"  [!] Không đọc được cookie nào từ {path}")
        return {}


def an_toan_ten_file(name: str) -> str:
    """Loại ký tự không hợp lệ trong tên file."""
    return re.sub(r'[\\/:*?"<>|]', "_", name)


def phat_su_kien(callback, su_kien: dict):
    """Gửi sự kiện cho callback (web UI) nếu có — CLI thì bỏ qua."""
    if callback:
        callback(su_kien)


def tai_anh(url: str, thumuc: str, ten_file: str) -> str | None:
    """Tải một ảnh về thư mục, trả về đường dẫn đã lưu (None nếu lỗi)."""
    try:
        resp = requests.get(url, headers={"User-Agent": DEFAULT_UA}, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"    [!] Lỗi tải ảnh {ten_file}: {e}")
        return None

    ext = ".jpg"
    ct = resp.headers.get("Content-Type", "")
    if "png" in ct:
        ext = ".png"
    elif "gif" in ct:
        ext = ".gif"
    elif "webp" in ct:
        ext = ".webp"

    path = os.path.join(thumuc, ten_file + ext)
    with open(path, "wb") as f:
        f.write(resp.content)
    return path


def xu_ly_post(post: dict, thumuc_anh: str | None, stt: int, tong: int,
               callback=None) -> dict:
    """Chuyển một post của facebook-scraper thành dict gọn, sạch và lưu ảnh."""
    post_id = post.get("post_id") or f"post_{stt}"
    text = (post.get("text") or "").strip()
    hashtags = re.findall(r"#[^\s#]+", text)

    # --- Ảnh: lấy tất cả link ảnh trong bài ---
    image_urls = list(dict.fromkeys(post.get("images") or []))

    anh_da_luu = []
    if thumuc_anh and image_urls:
        os.makedirs(thumuc_anh, exist_ok=True)
        for i, url in enumerate(image_urls):
            print(f"    [↓] Tải ảnh {i + 1}/{len(image_urls)} ...")
            # Tất cả ảnh nằm chung 1 thư mục (tên file có post_id để khỏi trùng)
            path = tai_anh(url, thumuc_anh,
                           f"{an_toan_ten_file(str(post_id))}_{i + 1}")
            if path:
                anh_da_luu.append(path)
            phat_su_kien(callback, {
                "loai": "anh", "post_id": post_id,
                "so_thu_tu": i + 1, "tong": len(image_urls),
                "duong_dan": path,
            })

    time_post = post.get("time")
    if isinstance(time_post, datetime):
        time_post = time_post.isoformat()

    print(
        f"  [✓] #{stt} | ID: {post_id} | {time_post} | "
        f"text {len(text)} ký tự | {len(image_urls)} ảnh | "
        f"{len(hashtags)} hashtag"
    )

    return {
        "post_id": post_id,
        "time": time_post,
        "text": text,
        "hashtags": hashtags,          # ví dụ: ['#PUBG', '#PUBGVN']
        "images": image_urls,          # link ảnh gốc trên Facebook
        "images_da_tai": anh_da_luu,   # đường dẫn ảnh đã tải về máy
        "post_url": post.get("post_url"),
        "likes": post.get("likes"),        # tổng số cảm xúc
        "comments": post.get("comments"),  # số bình luận
        "shares": post.get("shares"),      # số lượt chia sẻ
        "diem_tiem_nang": 0,               # điểm tương tác (tính sau)
        "muc_tiem_nang": "—",              # CAO / TRUNG_BINH / THAP (tính sau)
    }


def danh_gia_tiem_nang(posts: list):
    """Chấm điểm tương tác và xếp hạng tiềm năng cho từng bài.

    Công thức: diem = likes + comments*2 + shares*3
    (bình luận và chia sẻ nặng hơn like vì thể hiện tương tác sâu).
    Xếp hạng theo thứ hạng trong chính trang đó:
    - Top 25%  -> CAO
    - 25-60%   -> TRUNG_BINH
    - còn lại  -> THAP
    """
    if len(posts) < 5:
        return  # quá ít bài, không đủ cơ sở xếp hạng
    for p in posts:
        like = p.get("likes") or 0
        cmt = p.get("comments") or 0
        share = p.get("shares") or 0
        p["diem_tiem_nang"] = like + cmt * 2 + share * 3

    posts.sort(key=lambda p: p["diem_tiem_nang"], reverse=True)
    if not posts or posts[0]["diem_tiem_nang"] == 0:
        return  # tất cả đều 0 tương tác -> không xếp hạng

    n = len(posts)
    for i, p in enumerate(posts):
        phan_tram = (i + 1) / n
        if phan_tram <= 0.25:
            p["muc_tiem_nang"] = "CAO"
        elif phan_tram <= 0.60:
            p["muc_tiem_nang"] = "TRUNG_BINH"
        else:
            p["muc_tiem_nang"] = "THAP"


def in_top_tiem_nang(posts: list, top_n: int = 5):
    """In danh sách bài tiềm năng nhất ra màn hình."""
    print(f"\n  🏆 Top {min(top_n, len(posts))} bài TIỀM NĂNG nhất:")
    for i, p in enumerate(posts[:top_n]):
        print(
            f"    {i + 1}. [{p['muc_tiem_nang']}] điểm {p['diem_tiem_nang']} "
            f"(cảm xúc {p['likes']}, bình luận {p['comments']}, chia sẻ {p['shares']})"
        )
        if p["text"]:
            print(f"       {p['text'][:100]}{'...' if len(p['text']) > 100 else ''}")
        print(f"       {p['post_url']}")


def ghi_json(posts: list, ten_file: str):
    """Ghi file JSON lưu trữ dữ liệu thô."""
    json_path = f"{ten_file}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(posts, f, ensure_ascii=False, indent=2)
    return json_path


def ghi_excel(ket_qua: dict, ten_file: str) -> str:
    """Xuất TOÀN BỘ dữ liệu ra 1 file Excel .xlsx.

    - Sheet 'Tổng quan': thống kê từng trang + Top 10 bài tiềm năng nhất toàn bộ
    - 1 sheet cho mỗi trang: từng bài kèm cảm xúc, bình luận, chia sẻ,
      điểm tiềm năng (tô màu: CAO xanh lá / TRUNG_BINH vàng / THAP xám),
      đường dẫn ảnh đã tải, link bài bấm được.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cot = [
        ("STT", 5), ("post_id", 16), ("Thời gian", 20), ("Nội dung bài viết", 65),
        ("Cảm xúc", 11), ("Bình luận", 11), ("Chia sẻ", 10),
        ("Điểm tiềm năng", 13), ("Mức tiềm năng", 14),
        ("Ảnh đã tải (đường dẫn)", 50), ("Link bài viết", 16),
    ]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    tier_fill = {
        "CAO": PatternFill("solid", fgColor="C6EFCE"),
        "TRUNG_BINH": PatternFill("solid", fgColor="FFEB9C"),
        "THAP": PatternFill("solid", fgColor="F2F2F2"),
    }
    tier_font = {
        "CAO": Font(color="006100", bold=True),
        "TRUNG_BINH": Font(color="9C6500", bold=True),
        "THAP": Font(color="808080"),
    }

    def ke_dong_dau(ws, headers):
        """Tô header, kẻ khung, lọc tự động, đóng băng dòng đầu."""
        for c, (tieu_de, _) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=tieu_de)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).coordinate}"
        for c, (_, rong) in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + c)].width = rong

    def ghi_mot_bai(ws, r, post):
        ws.cell(r, 1, r - 1)
        ws.cell(r, 2, post.get("post_id"))
        ws.cell(r, 3, post.get("time"))
        # Hashtag được ghi NGAY TRONG ô nội dung, xuống dòng phía dưới bài viết
        text = post.get("text") or ""
        hashtags = " ".join(post.get("hashtags") or [])
        noi_dung = text + ("\n\n" + hashtags if hashtags else "")
        ws.cell(r, 4, noi_dung).alignment = Alignment(
            wrap_text=True, vertical="top")
        # Chiều cao dòng theo lượng chữ — nếu không đặt, Excel chỉ hiện 1 dòng
        # đầu của mỗi bài (text vẫn đầy đủ trong ô nhưng bị ẩn).
        # Cột rộng 65 -> ~60 ký tự/dòng. Excel giới hạn dòng 409.5pt.
        so_dong = max(1, -(-len(noi_dung) // 60))
        ws.row_dimensions[r].height = min(so_dong * 15 + 4, 409.5)
        ws.cell(r, 5, post.get("likes") or 0)
        ws.cell(r, 6, post.get("comments") or 0)
        ws.cell(r, 7, post.get("shares") or 0)
        ws.cell(r, 8, post.get("diem_tiem_nang") or 0)
        muc = post.get("muc_tiem_nang") or "—"
        c = ws.cell(r, 9, muc)
        if muc in tier_fill:
            c.fill = tier_fill[muc]
            c.font = tier_font[muc]
        # Đường dẫn ẢNH ĐẦY ĐỦ trên máy
        # (VD: C:\traffic fb\du_lieu_images\du_lieu_2026-08-24_09-27-08_images\post_id_1.jpg)
        cac_anh = [os.path.abspath(x) for x in (post.get("images_da_tai") or [])]
        ws.cell(r, 10, "\n".join(cac_anh)).alignment = Alignment(
            wrap_text=True, vertical="top")
        if cac_anh:
            # dòng cao thêm nếu nhiều ảnh
            ws.row_dimensions[r].height = max(ws.row_dimensions[r].height,
                                              min(len(cac_anh) * 15 + 4, 409.5))
        # Link bài viết: dán URL THẬT vào ô (bấm vẫn mở được Facebook)
        url = post.get("post_url")
        if url:
            c = ws.cell(r, 11, url)
            c.hyperlink = url
            c.font = Font(color="0563C1", underline="single")
        else:
            ws.cell(r, 11, "—")
        for c in range(1, len(cot) + 1):
            ws.cell(r, c).border = border

    wb = Workbook()

    # ================= Sheet Tổng quan =================
    ws = wb.active
    ws.title = "Tổng quan"
    ws["A1"] = "TỔNG QUAN CÁC TRANG ĐÃ CÀO"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Trang"; ws["B3"] = "Số bài"; ws["C3"] = "Tổng cảm xúc"
    ws["D3"] = "Tổng bình luận"; ws["E3"] = "Tổng chia sẻ"; ws["F3"] = "Tổng điểm"; ws["G3"] = "Bài hay nhất"
    for c in range(1, 8):
        ws.cell(3, c).fill = header_fill
        ws.cell(3, c).font = header_font
    ws.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15
    ws.column_dimensions["G"].width = 45

    r = 4
    tat_ca = []
    for page, posts in ket_qua.items():
        tong_diem = sum((p.get("diem_tiem_nang") or 0) for p in posts)
        bai_hay = posts[0] if posts and posts[0].get("post_url") else None
        ws.cell(r, 1, page)
        ws.cell(r, 2, len(posts))
        ws.cell(r, 3, sum(p.get("likes") or 0 for p in posts))
        ws.cell(r, 4, sum(p.get("comments") or 0 for p in posts))
        ws.cell(r, 5, sum(p.get("shares") or 0 for p in posts))
        ws.cell(r, 6, tong_diem)
        if bai_hay:
            c = ws.cell(r, 7, (bai_hay.get("text") or "")[:60] or bai_hay["post_id"])
            c.hyperlink = bai_hay["post_url"]
            c.font = Font(color="0563C1", underline="single")
        tat_ca.extend((p | {"_trang": page}) for p in posts)
        r += 1

    r += 1
    ws.cell(r, 1, "→ Cảm xúc / bình luận / chia sẻ CỦA TỪNG BÀI: mở sheet riêng của mỗi trang "
                  "(tab phía dưới, đặt tên theo trang)").font = Font(italic=True, color="64748B")
    r += 1
    ws.cell(r, 1, "TOP 10 BÀI TIỀM NĂNG NHẤT TOÀN BỘ").font = Font(bold=True, size=13)
    r += 1
    ws.cell(r, 1, "STT"); ws.cell(r, 2, "Trang"); ws.cell(r, 3, "post_id")
    ws.cell(r, 4, "Nội dung"); ws.cell(r, 5, "Điểm"); ws.cell(r, 6, "Mức"); ws.cell(r, 7, "Link")
    for c in range(1, 8):
        ws.cell(r, c).fill = header_fill
        ws.cell(r, c).font = header_font
    r += 1
    for i, p in enumerate(sorted(tat_ca, key=lambda x: x.get("diem_tiem_nang") or 0, reverse=True)[:10], 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, p.get("_trang"))
        ws.cell(r, 3, p.get("post_id"))
        ws.cell(r, 4, (p.get("text") or "")[:100])
        ws.cell(r, 5, p.get("diem_tiem_nang") or 0)
        muc = p.get("muc_tiem_nang") or "—"
        c = ws.cell(r, 6, muc)
        if muc in tier_fill:
            c.fill = tier_fill[muc]
            c.font = tier_font[muc]
        if p.get("post_url"):
            c = ws.cell(r, 7, "Mở bài viết")
            c.hyperlink = p["post_url"]
            c.font = Font(color="0563C1", underline="single")
        r += 1

    # ================= Sheet từng trang =================
    for page, posts in ket_qua.items():
        ten_sheet = re.sub(r'[\\/*?:\[\]]', "_", page)[:31] or "trang"
        ws = wb.create_sheet(title=ten_sheet)
        ke_dong_dau(ws, cot)
        for i, post in enumerate(posts, 2):
            ghi_mot_bai(ws, i, post)

    xlsx_path = f"{ten_file}.xlsx"
    thumuc = os.path.dirname(xlsx_path)
    if thumuc:
        os.makedirs(thumuc, exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


def cào_mot_trang(page: str, pages_scroll: int, limit: int, thumuc_anh: str | None,
                  delay: float, callback=None, stop_flag=None,
                  cookies_playwright: list = None) -> list:
    """Cào một trang, trả về danh sách post đã xử lý.

    - Có cookies -> dùng engine mới (cao_fb): mở trình duyệt thật, cuộn, đọc DOM.
    - Không có cookies -> thử thư viện facebook-scraper cũ (thường bị FB chặn).
    """
    posts = []
    phat_su_kien(callback, {"loai": "trang_bat_dau", "page": page, "gioi_han": limit})
    try:
        if cookies_playwright:
            bai_tho = cao_fb.cào_trang(
                page, so_bai=limit or 0, so_lan_cuon=pages_scroll,
                delay=delay, cookies=cookies_playwright,
                stop_flag=stop_flag, callback=callback,
            )
            for bai in bai_tho:
                if stop_flag and stop_flag.is_set():
                    print(f"  [i] Đã dừng theo yêu cầu tại trang {page}")
                    break
                post_da_xu_ly = xu_ly_post(bai, thumuc_anh, len(posts) + 1,
                                           limit or "?", callback=callback)
                posts.append(post_da_xu_ly)
                phat_su_kien(callback, {
                    "loai": "post", "page": page, "post": post_da_xu_ly,
                    "stt": len(posts), "tong": limit or None,
                })
                if limit and len(posts) >= limit:
                    break
                nghi_stop(delay, stop_flag)
        elif get_posts:
            for post in get_posts(
                page,
                pages=pages_scroll,
                options={"posts_per_page": 200},
            ):
                if stop_flag and stop_flag.is_set():
                    print(f"  [i] Đã dừng theo yêu cầu tại trang {page}")
                    break
                post_da_xu_ly = xu_ly_post(post, thumuc_anh, len(posts) + 1,
                                           limit or "?", callback=callback)
                posts.append(post_da_xu_ly)
                phat_su_kien(callback, {
                    "loai": "post", "page": page, "post": post_da_xu_ly,
                    "stt": len(posts), "tong": limit or None,
                })
                if limit and len(posts) >= limit:
                    break
                nghi_stop(delay, stop_flag)
        else:
            raise RuntimeError(
                "Thiếu cookies và thư viện cũ. Hãy xuất cookies.txt rồi chạy lại.")
    except Exception as e:
        print(f"\n[!] Lỗi khi cào: {e}")
        phat_su_kien(callback, {"loai": "loi", "page": page, "noi_dung": str(e)})
        raise
    phat_su_kien(callback, {"loai": "trang_xong", "page": page, "so_bai": len(posts)})
    return posts


def run_cào(trang_ho: list, pages: int = 3, limit: int = 0, per_page: int = 0,
            output: str = "du_lieu", no_images: bool = False, images_dir: str = None,
            cookies: str = None, delay: float = 3.0,
            callback=None, stop_flag=None):
    """Phần lõi cào — dùng chung cho CLI (main) và Web UI.

    Trả về (ket_qua, xlsx_path) với ket_qua = {trang: [posts]}.
    Mỗi bước đều gửi sự kiện qua callback nếu có.
    """
    print(f"=== Cào {len(trang_ho)} trang: {', '.join(trang_ho)} ===")
    print(f"    {pages} lần cuộn/trang, tối đa {per_page or limit or 'không'} bài/trang")
    cookies_playwright = cao_fb.doc_cookies(cookies) if cookies else None

    # 1 lần cào = 1 giờ cào — dùng chung cho cả thư mục ảnh và tên file Excel
    thoi_gian_cao = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    # TẤT CẢ ảnh của lần cào nằm trong MỘT thư mục đặt tên theo giờ cào,
    # và thư mục đó nằm BÊN TRONG du_lieu_images/
    # (VD: du_lieu_images/du_lieu_2026-08-24_09-52-42_images/)
    thumuc_anh = None if no_images else (
        images_dir or os.path.join("du_lieu_images",
                                   f"{output}_{thoi_gian_cao}_images"))

    ket_qua = {}
    for i, page in enumerate(trang_ho):
        print(f"\n----- Trang {i + 1}/{len(trang_ho)}: {page} -----")
        page_id_sach = an_toan_ten_file(in_tu_page(page))
        ten_file = output if len(trang_ho) == 1 else f"{output}_{page_id_sach}"

        posts = cào_mot_trang(page, pages, per_page or limit, thumuc_anh, delay,
                              callback=callback, stop_flag=stop_flag,
                              cookies_playwright=cookies_playwright)

        if posts:
            danh_gia_tiem_nang(posts)
            in_top_tiem_nang(posts)
        else:
            phat_su_kien(callback, {"loai": "khong_bai", "page": page})

        json_path = ghi_json(posts, ten_file)
        print(f"  ✅ {page}: {len(posts)} bài -> {json_path}")
        if thumuc_anh:
            print(f"     Ảnh: {thumuc_anh}")
        ket_qua[page] = posts

        if i < len(trang_ho) - 1 and delay > 0:
            print(f"  ... nghỉ {delay}s trước khi sang trang tiếp theo")
            nghi_stop(delay, stop_flag)

    # Mỗi lần chạy = 1 file Excel RIÊNG trong folder 'du_lieu_exel'
    # (kèm giờ phút để không ghi đè lần chạy trước — cùng giờ với thư mục ảnh)
    ten_xlsx = f"du_lieu_exel/{output}_{thoi_gian_cao}"
    xlsx_path = ghi_excel(ket_qua, ten_xlsx)

    print(f"\n=== HOÀN THÀNH: {sum(len(p) for p in ket_qua.values())} bài viết từ {len(ket_qua)} trang ===")
    for page, posts in ket_qua.items():
        print(f"  - {page}: {len(posts)} bài")
    print(f"\n📊 MỞ FILE EXCEL: {xlsx_path}")
    print("   (sheet 'Tổng quan' = thống kê + top 10 tiềm năng; 1 sheet/trang)")
    return ket_qua, xlsx_path


def main():
    parser = argparse.ArgumentParser(
        description="Cào bài viết + ảnh từ nhiều Facebook Page (trang public)."
    )
    parser.add_argument(
        "--page", nargs="+", required=True,
        help="Tên hoặc URL trang — ghi được NHIỀU trang, cách nhau bằng dấu cách",
    )
    parser.add_argument(
        "--pages", type=int, default=3,
        help="Số lần cuộn cho mỗi trang (mỗi lần ~200 bài, mặc định 3)",
    )
    parser.add_argument(
        "--limit", type=int, default=0,
        help="Giới hạn số bài (0 = không giới hạn)",
    )
    parser.add_argument(
        "--per-page", type=int, default=0,
        help="Giới hạn số bài cho MỖI trang (0 = dùng --limit)",
    )
    parser.add_argument(
        "--output", default="du_lieu",
        help="Tiền tố tên file đầu ra, mặc định: du_lieu (file: du_lieu_<trang>.json + du_lieu.xlsx)",
    )
    parser.add_argument(
        "--images-dir", default=None,
        help="Thư mục lưu ảnh (mặc định: du_lieu_images/<tiền tố>_<giờ-cào>_images — tất cả ảnh 1 lần cào)",
    )
    parser.add_argument(
        "--no-images", action="store_true",
        help="Không tải ảnh, chỉ lấy link ảnh",
    )
    parser.add_argument(
        "--cookies", default=None,
        help="File cookies.txt (bắt buộc vì FB chặn truy cập ẩn danh)",
    )
    parser.add_argument(
        "--delay", type=float, default=3.0,
        help="Giây nghỉ giữa mỗi lần cuộn và giữa các trang (giảm rủi ro bị chặn)",
    )
    args = parser.parse_args()

    trang_ho = [in_tu_page(p) for p in args.page]
    ket_qua, xlsx_path = run_cào(
        trang_ho, pages=args.pages, limit=args.limit, per_page=args.per_page,
        output=args.output, no_images=args.no_images, images_dir=args.images_dir,
        cookies=args.cookies, delay=args.delay,
    )

    if not any(ket_qua.values()):
        print("\n[!] Không lấy được bài nào. Facebook hiện CHẶN truy cập ẩn danh (không đăng nhập).")
        print("    Giải pháp: xuất cookies từ trình duyệt đang đăng nhập Facebook rồi chạy lại:\n")
        print("    Cách 1 - Dùng extension (dễ nhất):")
        print("        Cài 'Get cookies.txt LOCALLY' cho Chrome/Edge (miễn phí trên store),")
        print("        mở trang facebook.com -> bấm icon extension -> Export -> lưu cookies.txt")
        print("    Cách 2 - Tự copy từ DevTools:")
        print("        Mở facebook.com (đã đăng nhập) -> F12 -> Console, dán lệnh:")
        print("        copy(document.cookie) -> lưu vào cookies.txt dạng:  ten_cookie\tgia_tri  (mỗi dòng 1 cặp)\n")
        print(f"    Sau đó chạy:")
        print(f"        python scraper.py --page {' '.join(trang_ho)} --cookies cookies.txt")
        sys.exit(1)


if __name__ == "__main__":
    main()
